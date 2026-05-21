import sys
import socket
# Fix Windows console encoding for emoji/Unicode characters
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

# Patch getaddrinfo to bypass DNS issues with supabase
_orig_getaddrinfo = socket.getaddrinfo
def patched_getaddrinfo(*args, **kwargs):
    try:
        host = args[0]
        if isinstance(host, str) and host.endswith('.supabase.co'):
            return _orig_getaddrinfo('104.18.38.10', *args[1:], **kwargs)
    except Exception:
        pass
    return _orig_getaddrinfo(*args, **kwargs)
socket.getaddrinfo = patched_getaddrinfo

from fastapi import FastAPI, HTTPException, Request, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.middleware.cors import CORSMiddleware
import secrets
from datetime import datetime, timedelta
from typing import Dict, List, Optional
import io
import os
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# Optional Twilio import (graceful if not installed)
try:
    from twilio.rest import Client as TwilioClient
    TWILIO_AVAILABLE = True
except ImportError:
    TwilioClient = None
    TWILIO_AVAILABLE = False
    print("[WARN] Twilio not installed - SMS OTP will be skipped")

from backend.models import (
    VoterRegistration, VoterLogin, AdminLogin, 
    ElectionCreate, VoteRequest, VoteVerification, Candidate, INDIAN_STATES,
    VoterFetchPhoneRequest, VoterRequestOtpRequest, VoterVerifyOtpRequest, VoterSignupOtpRequest
)
from backend.auth import VoterAuth, AdminAuth, MockAadhaarAuth
from backend.db_config import get_database
from backend.encryption import hash_voter_token

app = FastAPI(
    title="SecureVoteChain - Blockchain E-Voting System",
    description="A secure, transparent blockchain-based electronic voting platform with Zero-Knowledge Proofs and Ring Signatures.",
    version="1.0.0"
)

# CORS Configuration - Allow localhost on all common ports + 127.0.0.1
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://localhost:5000",
        "http://localhost:8000",
        "http://127.0.0.1:3000",
        "http://127.0.0.1:5000",
        "http://127.0.0.1:8000",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

db = get_database()

# Configure AdminAuth to use database for authentication
AdminAuth.set_database(db)
print("[OK] AdminAuth configured to use database authentication")

templates = Jinja2Templates(directory="templates")
app.mount("/static", StaticFiles(directory="static"), name="static")


# ─── Startup Event: Auto-update election statuses ─────────────────────────────
@app.on_event("startup")
async def startup_tasks():
    """Run background tasks on server startup."""
    try:
        _auto_update_election_statuses()
        print("[OK] Startup tasks completed")
    except Exception as e:
        print(f"[WARN] Startup tasks failed: {e}")


def _auto_update_election_statuses():
    """Mark elections as 'ended' if their end_time has passed."""
    try:
        elections = db.get_all_elections()
        now = datetime.now()
        updated = 0
        for election in elections:
            if election.get("status") == "active":
                end_time_str = election.get("end_time", "")
                if end_time_str:
                    try:
                        end_time = datetime.fromisoformat(end_time_str.replace("Z", ""))
                        if now > end_time:
                            if hasattr(db, 'update_election_status'):
                                db.update_election_status(election["id"], "ended")
                            updated += 1
                    except Exception:
                        pass
        if updated:
            print(f"[OK] Auto-closed {updated} expired election(s)")
    except Exception as e:
        print(f"[WARN] Could not auto-update election statuses: {e}")


# Middleware to check role-based access
def check_admin_access(request: Request) -> Dict:
    """Verify admin authentication and return session data."""
    auth_header = request.headers.get("Authorization")
    if not auth_header:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    session = db.get_session(auth_header.replace("Bearer ", ""))
    if not session or session.get("type") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    return session


def check_voter_access(request: Request) -> Dict:
    """Verify voter authentication and return session data."""
    auth_header = request.headers.get("Authorization")
    if not auth_header:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    session = db.get_session(auth_header.replace("Bearer ", ""))
    if not session or session.get("type") != "voter":
        raise HTTPException(status_code=403, detail="Voter access required")
    
    return session


def normalize_phone_number(phone: Optional[str]) -> Optional[str]:
    if not phone:
        return None

    cleaned = ''.join(ch for ch in str(phone).strip() if ch.isdigit() or ch == '+')
    if not cleaned:
        return None

    if cleaned.startswith('+'):
        return cleaned

    if cleaned.startswith('91') and len(cleaned) >= 12:
        return '+' + cleaned

    if cleaned.startswith('0') and len(cleaned) >= 11:
        cleaned = cleaned[-10:]

    if len(cleaned) == 10:
        return '+91' + cleaned

    return '+91' + cleaned


def get_vote_tracking_records() -> List[Dict]:
    """Get vote tracking records, preferring Supabase when available."""
    vote_tracking = []
    try:
        if hasattr(db, 'client'):
            page_size = 1000
            offset = 0
            while True:
                result = db.client.table('vote_tracking').select('*').range(offset, offset + page_size - 1).execute()
                batch = [dict(row) for row in result.data] if result.data else []

                if not batch:
                    break

                vote_tracking.extend(batch)

                if len(batch) < page_size:
                    break

                offset += page_size
    except Exception as e:
        print(f"⚠️ Error getting vote tracking: {e}")

    if not vote_tracking:
        all_votes = db.get_all_votes()
        for vote in all_votes:
            election_id = vote.get('election_id')
            voter_identifier = vote.get('voter_token') or vote.get('voter_token_hash')
            if election_id and voter_identifier:
                vote_tracking.append({
                    'election_id': election_id,
                    'voter_token_hash': voter_identifier
                })

    return vote_tracking


def normalize_audit_log_entry(log: Dict) -> Dict:
    """Normalize audit log data for the admin UI."""
    action_type = log.get('action_type') or log.get('action') or 'unknown'
    action_details = log.get('action_details') or log.get('details') or ''

    return {
        'id': log.get('id'),
        'timestamp': log.get('timestamp'),
        'action_type': action_type,
        'action': action_type,
        'action_details': action_details,
        'details': action_details,
        'admin_username': log.get('admin_username') or log.get('username') or log.get('user_id') or 'unknown',
        'admin_state': log.get('admin_state') or log.get('state') or 'N/A',
        'admin_role': log.get('admin_role') or log.get('role') or 'admin',
        'ip_address': log.get('ip_address') or 'N/A'
    }


def audit_log_matches_filters(log: Dict, filters: Dict[str, Optional[str]]) -> bool:
    """Return True when an audit log matches the supplied filters."""
    search = (filters.get('search') or '').strip().lower()
    action = (filters.get('action') or '').strip().lower()
    username = (filters.get('username') or '').strip().lower()
    state = (filters.get('state') or '').strip().lower()
    role = (filters.get('role') or '').strip().lower()
    from_date = filters.get('from_date')
    to_date = filters.get('to_date')

    log_action = str(log.get('action_type') or log.get('action') or '').lower()
    log_username = str(log.get('admin_username') or log.get('username') or log.get('user_id') or '').lower()
    log_state = str(log.get('admin_state') or log.get('state') or '').lower()
    log_role = str(log.get('admin_role') or log.get('role') or '').lower()
    log_details = str(log.get('action_details') or log.get('details') or '').lower()
    log_text = ' '.join([log_action, log_username, log_state, log_role, log_details]).strip()

    if action and action not in log_action:
        return False
    if username and username not in log_username:
        return False
    if state and state not in log_state:
        return False
    if role and role not in log_role:
        return False
    if search and search not in log_text:
        return False

    if from_date or to_date:
        try:
            log_timestamp = datetime.fromisoformat(str(log.get('timestamp', '')).replace('Z', '+00:00'))
            if from_date:
                start_date = datetime.fromisoformat(from_date)
                if log_timestamp < start_date:
                    return False
            if to_date:
                end_date = datetime.fromisoformat(to_date)
                if log_timestamp > end_date:
                    return False
        except Exception:
            pass

    return True


@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    """Landing page."""
    return templates.TemplateResponse(request=request, name="index.html", context={})


@app.get("/admin", response_class=HTMLResponse)
async def admin_page(request: Request):
    """Admin dashboard - Only for admins."""
    # Check if there's a session token in query params or cookies
    auth_header = request.headers.get("Authorization", "")
    if auth_header:
        session = db.get_session(auth_header.replace("Bearer ", ""))
        if session and session.get("type") == "voter":
            # Voter trying to access admin panel - redirect to voter page
            from fastapi.responses import RedirectResponse
            return RedirectResponse(url="/voter")
    return templates.TemplateResponse(request=request, name="admin.html", context={})


@app.get("/voter", response_class=HTMLResponse)
async def voter_page(request: Request):
    """Voter interface - Only for voters."""
    # Check if there's a session token in query params or cookies
    auth_header = request.headers.get("Authorization", "")
    if auth_header:
        session = db.get_session(auth_header.replace("Bearer ", ""))
        if session and session.get("type") == "admin":
            # Admin trying to access voter panel - redirect to admin page
            from fastapi.responses import RedirectResponse
            return RedirectResponse(url="/admin")
    return templates.TemplateResponse(request=request, name="voter.html", context={})


@app.get("/verify", response_class=HTMLResponse)
async def verify_page(request: Request):
    """Public vote verification portal."""
    return templates.TemplateResponse(request=request, name="verify.html", context={})


@app.get("/statistics", response_class=HTMLResponse)
async def statistics_page(request: Request):
    """Public statistics dashboard showing voting participation."""
    return templates.TemplateResponse(request=request, name="statistics.html", context={})


@app.get("/candidate", response_class=HTMLResponse)
async def candidate_profile_page(request: Request):
    """Candidate profile page with detailed information."""
    return templates.TemplateResponse(request=request, name="candidate.html", context={})


@app.post("/api/admin/login")
async def admin_login(login: AdminLogin):
    """Admin login endpoint with state-based access."""
    try:
        success, error, admin_data = AdminAuth.authenticate_admin(login.username, login.password)
        
        if not success or not admin_data:
            print(f"❌ Authentication failed for user: {login.username}")
            raise HTTPException(status_code=401, detail=error or "Authentication failed")
        
        session_token = AdminAuth.generate_admin_session_token()
        print(f"✅ Admin authenticated: {login.username} ({admin_data['state']})")
        
        # Save session with correct parameters: (session_token, user_id, session_type, user_data)
        db.save_session(session_token, admin_data["username"], "admin", {
            "username": admin_data["username"],
            "state": admin_data["state"],
            "role": admin_data["role"],
            "created_at": datetime.now().isoformat()
        })
        
        print(f"✅ Session saved for {login.username}")
        
        return {
            "success": True,
            "session_token": session_token,
            "username": admin_data["username"],
            "state": admin_data["state"],
            "role": admin_data["role"]
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Login error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Server error: {str(e)}")


def send_otp_sms(phone, otp_code, context):
    try:
        if not TWILIO_AVAILABLE or TwilioClient is None:
            print("[WARN] Twilio not available - SMS skipped")
            return False
        twilio_sid = os.getenv("TWILIO_ACCOUNT_SID")
        twilio_token = os.getenv("TWILIO_AUTH_TOKEN")
        twilio_from_number = os.getenv("TWILIO_PHONE_NUMBER") or os.getenv("TWILIO_FROM_NUMBER") or os.getenv("TWILIO_MESSAGING_NUMBER")
        
        if twilio_sid and twilio_token and twilio_from_number:
            client = TwilioClient(twilio_sid, twilio_token)
            message = client.messages.create(
                body=f"SecureVoteChain OTP [{context}]: {otp_code}. Valid for 5 minutes. Do not share.",
                from_=twilio_from_number,
                to=phone,
            )
            print(f"[OK] Twilio SMS sent successfully to {phone} (SID: {message.sid})")
            return True
        else:
            print("[WARN] Twilio SMS credentials missing from .env. Skipping SMS.")
            return False
    except Exception as twilio_error:
        print(f"❌ Twilio SMS failed: {twilio_error}")
        return str(twilio_error)


@app.post("/api/voter/fetch-phone")
async def fetch_phone(data: VoterFetchPhoneRequest):
    voter_id = data.voter_id
    aadhaar_number = data.aadhaar_number
    
    if not aadhaar_number or len(aadhaar_number) != 12:
        raise HTTPException(status_code=400, detail="Invalid inputs")
        
    voter = db.get_voter_by_aadhaar(aadhaar_number)
    if not voter or (voter_id and voter.get("voter_id") != voter_id):
        raise HTTPException(status_code=404, detail="Invalid Voter ID or Aadhaar Number")
        
    phone = db.get_phone_by_aadhaar(aadhaar_number)
    if phone:
        masked_phone = phone[:3] + "******" + phone[-3:] if len(phone) > 10 else "***" + phone[-4:]
        return {"success": True, "phone": masked_phone}
    else:
         raise HTTPException(status_code=404, detail="No registered mobile number found")

@app.post("/api/voter/request-otp")
async def request_otp(data: VoterRequestOtpRequest):
    """
    Request OTP for voter login via SMS.
    """
    aadhaar_number = data.aadhaar_number
    voter_id = data.voter_id
    
    if not aadhaar_number or len(aadhaar_number) != 12:
        raise HTTPException(status_code=400, detail="Invalid Aadhaar number (must be 12 digits)")
    
    voter = db.get_voter_by_aadhaar(aadhaar_number)
    if not voter or (voter_id and voter.get("voter_id") != voter_id):
        raise HTTPException(status_code=404, detail="Aadhaar number not registered or Voter ID mismatch.")
        
    phone = db.get_phone_by_aadhaar(aadhaar_number)

    phone = normalize_phone_number(phone)
    
    # Generate and store OTP (6 digits)
    otp_code = ''.join([str(secrets.randbelow(10)) for _ in range(6)])
    
    # Store OTP in session with expiration (5 minutes)
    otp_session_key = f"otp_{aadhaar_number}"
    
    # Delete any existing OTP session for this Aadhaar
    try:
        db.delete_session(otp_session_key)
    except:
        pass  # Ignore if session doesn't exist
    
    # Save new OTP session (5 min expiry)
    db.save_session(otp_session_key, aadhaar_number, "otp", {
        "otp": otp_code,
        "phone": phone,
        "created_at": datetime.now().isoformat(),
        "expires_at": (datetime.now() + timedelta(minutes=5)).isoformat()
    })
    
    # ALWAYS show OTP in terminal for easy access
    print(f"\n{'='*60}")
    print(f"🔐 OTP GENERATED FOR AADHAAR: {aadhaar_number[:4]}****{aadhaar_number[-4:]}")
    print(f"📱 Phone: {phone}")
    print(f"🔢 OTP CODE: {otp_code}")
    print(f"⏰ Valid for: 5 minutes")
    print(f"{'='*60}\n")
    
    # Send OTP via Twilio SMS, then fall back to terminal logging in local mode.
    has_phone = bool(phone)
    sms_sent = False
    
    if has_phone:
        sms_sent = send_otp_sms(phone, otp_code, "Login")
    else:
        print("⚠️ No phone found for voter. Using terminal OTP for local login.")
        print(f"💡 Use the OTP shown above in terminal: {otp_code}")
    
    # Mask phone for privacy (show last 4 digits)
    masked_phone = "***" + phone[-4:] if phone and len(phone) >= 4 else "local-mode"
    
    if sms_sent is True:
        final_msg = f"OTP sent to {masked_phone} via SMS. Check your messages."
    elif isinstance(sms_sent, str):
        final_msg = f"SMS failed: {sms_sent}"
    else:
        final_msg = f"OTP generated for {masked_phone}, but SMS delivery is not configured. Check the server terminal."

    return {
        "success": True,
        "sms_sent": sms_sent is True,
        "message": final_msg,
        "phone_masked": masked_phone,
        "otp_hint": f"OTP starts with {otp_code[:2]}**",
        "expires_in": "5 minutes"
    }


@app.post("/api/voter/verify-otp")
async def verify_otp(data: VoterVerifyOtpRequest):
    """
    Verify OTP and log in voter.
    """
    voter_id = data.voter_id
    aadhaar_number = data.aadhaar_number
    otp_code = data.otp
    
    if not aadhaar_number or len(aadhaar_number) != 12:
        raise HTTPException(status_code=400, detail="Invalid Aadhaar number")
    
    if not otp_code or len(otp_code) != 6:
        raise HTTPException(status_code=400, detail="Invalid OTP (must be 6 digits)")
    
    # Get voter from database
    voter = db.get_voter_by_aadhaar(aadhaar_number)
    
    if not voter or (voter_id and voter.get("voter_id") != voter_id):
        raise HTTPException(status_code=404, detail="Voter not found or mismatch.")
    
    email = voter.get("email") or f"{voter['voter_id'].lower()}@local.test"
    
    # Verify OTP from stored session
    otp_session_key = f"otp_{aadhaar_number}"
    otp_session = db.get_session(otp_session_key)
    
    if not otp_session:
        raise HTTPException(status_code=401, detail="OTP expired or not found. Please request a new OTP.")
    
    # Check if OTP is expired (5 minutes)
    expires_at = datetime.fromisoformat(otp_session.get("expires_at", ""))
    if datetime.now() > expires_at:
        # Clean up expired OTP
        db.delete_session(otp_session_key)
        raise HTTPException(status_code=401, detail="OTP expired after 5 minutes. Please request a new OTP.")
    
    # Verify OTP matches
    stored_otp = otp_session.get("otp")
    if stored_otp != otp_code:
            raise HTTPException(status_code=401, detail="Invalid OTP. Please check and try again.")
    
    # OTP is valid - delete it (single use)
    db.delete_session(otp_session_key)
    
    # Create session token
    session_token = secrets.token_hex(32)
    db.save_session(session_token, voter["voter_id"], "voter", {
        "voter_id": voter["voter_id"],
        "voter_token": voter.get("voting_token", ""),
        "name": voter["name"],
        "state": voter.get("state", "Not specified"),
        "email": email,
        "created_at": datetime.now().isoformat()
    })
    
    print(f"✅ Voter logged in successfully: {voter['voter_id']}")
    
    return {
        "success": True,
        "voter_id": voter["voter_id"],
        "voter_token": voter.get("voting_token", ""),
        "name": voter["name"],
        "state": voter.get("state", "Not specified"),
        "session_token": session_token,
        "message": "Login successful!"
    }


@app.post("/api/voter/request-otp-signup")
async def request_otp_signup(data: VoterSignupOtpRequest):
    aadhaar_number = data.aadhaar_number
    phone = data.phone
    name = data.name
    
    if not aadhaar_number or len(aadhaar_number) != 12:
        raise HTTPException(status_code=400, detail="Invalid Aadhaar number")
    if not phone or len(phone) < 10:
        raise HTTPException(status_code=400, detail="Invalid phone number")
        
    phone = normalize_phone_number(phone)
    
    # Check if voter already exists
    voter = db.get_voter_by_aadhaar(aadhaar_number)
    if voter:
        raise HTTPException(
            status_code=400, 
            detail="A voter with this Aadhaar number is already registered. Please login using your Voter ID instead."
        )
    
    # Check if phone number is already used
    if hasattr(db, 'get_phone_by_aadhaar'):
        # Just sanity-checking if phone exists for another voter, but we don't have a direct phone lookup by phone in the wrapper easily.
        # It's better to just block the Aadhar first.
        pass

    otp_code = ''.join([str(secrets.randbelow(10)) for _ in range(6)])
    otp_session_key = f"otp_reg_{aadhaar_number}"
    
    try:
        db.delete_session(otp_session_key)
    except:
        pass
        
    db.save_session(otp_session_key, aadhaar_number, "otp_reg", {
        "otp": otp_code,
        "phone": phone,
        "name": name,
        "created_at": datetime.now().isoformat(),
        "expires_at": (datetime.now() + timedelta(minutes=5)).isoformat()
    })
    
    print(f"\n{'='*60}")
    print(f"?? REGISTRATION OTP GENERATED FOR AADHAAR: {aadhaar_number[:4]}****{aadhaar_number[-4:]}")
    print(f"?? Phone: {phone}")
    print(f"?? OTP CODE: {otp_code}")
    print(f"? Valid for: 5 minutes")
    print(f"{'='*60}\n")
    
    # Send registration OTP through the same SMS path used for login.
    sms_sent = send_otp_sms(phone, otp_code, "Registration")
    
    masked_phone = "***" + phone[-4:] if len(phone) >= 4 else "local-mode"
    
    if sms_sent is True:
        final_msg = f"OTP sent to {masked_phone} via SMS. Check your messages."
    elif isinstance(sms_sent, str):
        final_msg = f"SMS failed: {sms_sent}"
    else:
        final_msg = f"OTP generated for {masked_phone}, but SMS delivery is not configured. Check the server terminal."

    return {
        "success": True,
        "sms_sent": sms_sent is True,
        "message": final_msg,
        "phone_masked": masked_phone,
        "otp_hint": f"OTP starts with {otp_code[:2]}**",
        "expires_in": "5 minutes"
    }

@app.post("/api/voter/register")
async def register_voter(registration: VoterRegistration):
    """Complete voter registration after OTP verification."""
    
    otp_session_key = f"otp_reg_{registration.aadhaar_number}"
    otp_session = db.get_session(otp_session_key)
    
    if not otp_session:
        raise HTTPException(status_code=401, detail="OTP expired or not found. Please request a new OTP.")
    
    expires_at = datetime.fromisoformat(otp_session.get("expires_at", ""))
    if datetime.now() > expires_at:
        db.delete_session(otp_session_key)
        raise HTTPException(status_code=401, detail="OTP expired after 5 minutes. Please request a new OTP.")
        
    if otp_session.get("otp") == registration.otp:
        pass # Valid local terminal OTP
    else:
        import os
        from twilio.rest import Client
        twilio_verify_sid = os.getenv("TWILIO_VERIFY_SERVICE_SID")
        if twilio_verify_sid:
            try:
                client = Client(os.getenv("TWILIO_ACCOUNT_SID"), os.getenv("TWILIO_AUTH_TOKEN"))
                phone_to_verify = otp_session.get("phone")
                if not phone_to_verify:
                     phone_to_verify = registration.phone
                     if not phone_to_verify.startswith("+"): phone_to_verify = "+91" + phone_to_verify[-10:]
                verification_check = client.verify.v2.services(twilio_verify_sid).verification_checks.create(to=phone_to_verify, code=registration.otp)
                if verification_check.status != "approved":
                    raise HTTPException(status_code=401, detail="Invalid OTP. Please check and try again.")
            except Exception as e:
                if isinstance(e, HTTPException): raise e
                raise HTTPException(status_code=401, detail=f"Invalid Twilio OTP or verification failed.")
        else:
            raise HTTPException(status_code=401, detail="Invalid OTP. Please check and try again.")
        
    db.delete_session(otp_session_key)
    
    # Generate keys and insert!
    voter_id = f"VOT{secrets.token_hex(4).upper()}"
    voting_token = secrets.token_hex(32)
    
    phone = registration.phone
    if not phone.startswith("+"):
        phone = "+91" + phone if phone.startswith("91") else "+91" + phone
        
    voter_data = {
        'voter_id': voter_id,
        'voting_token': voting_token,
        'name': registration.name or otp_session.get("name"),
        'aadhaar': registration.aadhaar_number,
        'state': registration.state,
        'phone': phone,
        'registered_at': datetime.now().isoformat()
    }
    
    # Double check no duplicate Aadhar before DB insert
    existing_voter = db.get_voter_by_aadhaar(registration.aadhaar_number)
    if existing_voter:
        raise HTTPException(
            status_code=400, 
            detail="A voter with this Aadhaar number is already registered. Please login using your Voter ID instead."
        )
    
    try:
        result = db.register_voter(voter_data)
        if result is False or (isinstance(result, dict) and not result.get("success", True)):
            # If voter already existed, just return their existing token or error
            existing_voter = db.get_voter_by_aadhaar(registration.aadhaar_number)
            if existing_voter:
                session_token = secrets.token_hex(32)
                db.save_session(session_token, existing_voter["voter_id"], "voter", {
                    "voter_id": existing_voter["voter_id"],
                    "voter_token": existing_voter["voting_token"],
                    "name": existing_voter["name"],
                    "state": existing_voter.get("state", "Not specified"),
                    "created_at": datetime.now().isoformat()
                })
                return {
                    "success": True,
                    "already_registered": True,
                    "voter_id": existing_voter["voter_id"],
                    "voter_token": existing_voter["voting_token"],
                    "name": existing_voter["name"],
                    "state": existing_voter.get("state", "Not specified"),
                    "session_token": session_token
                }
            raise HTTPException(status_code=400, detail="Database registration failed")
            
    except Exception as e:
        print(f"Error saving voter to database: {e}")
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    
    session_token = secrets.token_hex(32)
    db.save_session(session_token, voter_data["voter_id"], "voter", {
        "voter_id": voter_data["voter_id"],
        "voter_token": voter_data["voting_token"],
        "name": voter_data["name"],
        "state": voter_data["state"],
        "created_at": datetime.now().isoformat()
    })

    return {
        "success": True,
        "voter_id": voter_data["voter_id"],
        "voter_token": voter_data["voting_token"],
        "name": voter_data["name"],
        "state": voter_data["state"],
        "session_token": session_token,
        "message": "Registration successful! Save your Voter ID and Token securely."
    }


@app.post("/api/voter/login")
async def voter_login(login: VoterLogin):
    """Voter login endpoint."""
    voter = db.get_voter(login.voter_id)
    
    if not voter:
        raise HTTPException(status_code=404, detail="Voter ID not found")
    
    session_token = secrets.token_hex(32)
    # Save session with correct parameters: (session_token, user_id, session_type, user_data)
    db.save_session(session_token, login.voter_id, "voter", {
        "voter_id": login.voter_id,
        "voter_token": voter["voting_token"],
        "name": voter["name"],
        "state": voter.get("state", "Not specified"),
        "created_at": datetime.now().isoformat()
    })
    
    return {
        "success": True,
        "session_token": session_token,
        "voter_id": login.voter_id,
        "name": voter["name"],
        "state": voter.get("state", "Not specified"),
        "voter_token": voter["voting_token"]
    }


@app.post("/api/admin/elections")
async def create_election(election_data: ElectionCreate, request: Request):
    """Create a new election (admin only) for specific state."""
    session = check_admin_access(request)
    
    # Verify admin can create elections for this state
    admin_state = session.get("state")
    if admin_state != "All States" and admin_state != election_data.state:
        raise HTTPException(
            status_code=403, 
            detail=f"You can only create elections for {admin_state}"
        )
    
    election_id = secrets.token_hex(8)
    
    election = {
        "id": election_id,
        "title": election_data.title,
        "description": election_data.description,
        "state": election_data.state,
        "start_time": election_data.start_time,
        "end_time": election_data.end_time,
        "candidates": [c.dict() for c in election_data.candidates],
        "status": "active",
        "created_at": datetime.now().isoformat(),
        "created_by": session.get("username")
    }
    
    db.save_election(election)
    
    return {
        "success": True,
        "election_id": election_id,
        "state": election_data.state,
        "message": f"Election created successfully for {election_data.state}"
    }


@app.get("/api/elections")
async def get_elections(request: Request):
    """Get elections filtered by user's state."""
    # Try to get session to determine user type and state
    auth_header = request.headers.get("Authorization")
    
    if auth_header:
        session = db.get_session(auth_header.replace("Bearer ", ""))
        if session:
            user_state = session.get("state")
            user_type = session.get("type")
            
            # Get all elections
            all_elections = db.get_all_elections()
            
            # Filter by state if not "All States" (for state admins and voters)
            if user_state and user_state != "All States":
                elections = [e for e in all_elections if e.get("state") == user_state]
            else:
                # Super admin sees all elections
                elections = all_elections
            
            return {
                "elections": elections, 
                "state": user_state,
                "is_super_admin": user_state == "All States"
            }
    
    # No auth - return empty or public elections only
    return {"elections": [], "state": None, "is_super_admin": False}


@app.get("/api/states")
async def get_states():
    """Get list of Indian states for dropdown."""
    return {"states": INDIAN_STATES}


@app.get("/api/elections/{election_id}")
async def get_election(election_id: str):
    """Get election details."""
    election = db.get_election_by_id(election_id)
    
    if not election:
        raise HTTPException(status_code=404, detail="Election not found")
    
    return election


@app.get("/api/admin/dashboard")
async def get_dashboard_stats(request: Request):
    """Get dashboard statistics for admin."""
    session = check_admin_access(request)
    
    elections = db.get_all_elections()
    voters = db.get_all_voters()
    
    # Filter by state if not super admin
    admin_state = session.get("state")
    if admin_state != "All States":
        elections = [e for e in elections if e.get("state") == admin_state]
        voters = [v for v in voters if v.get("state") == admin_state]
    
    total_votes = sum(len(db.get_votes_by_election(e["id"])) for e in elections)
    
    return {
        "elections": len(elections),
        "voters": len(voters),
        "votes": total_votes,
        "admin_state": admin_state,
        "is_super_admin": admin_state == "All States"
    }


@app.get("/api/admin/voters")
async def get_voters(request: Request):
    """Get voters filtered by admin's state."""
    session = check_admin_access(request)
    
    all_voters = db.get_all_voters()
    
    # Filter by state if not super admin
    admin_state = session.get("state")
    if admin_state != "All States":
        voters = [v for v in all_voters if v.get("state") == admin_state]
    else:
        voters = all_voters
    
    return {
        "voters": voters,
        "total": len(voters),
        "admin_state": admin_state
    }


@app.post("/api/admin/import-voters")
async def import_voters(request: Request, file: UploadFile = File(...)):
    """Import voters from CSV file (admin only)."""
    session = check_admin_access(request)
    admin_state = session.get("state")
    
    try:
        contents = await file.read()

        import csv
        import io
        import secrets

        filename = (file.filename or "").lower()
        voter_rows = []

        def normalize_header(value):
            return str(value or "").strip().lower().replace(" ", "_").replace("-", "_")

        def normalize_row(row):
            normalized = {}
            for key, value in row.items():
                normalized_key = normalize_header(key)
                if normalized_key:
                    normalized[normalized_key] = value

            aliases = {
                "aadhaar_number": "aadhaar",
                "aadhaarno": "aadhaar",
                "aadhar_number": "aadhaar",
                "aadhar": "aadhaar",
                "phone_number": "phone",
                "mobile_number": "phone",
                "mobile": "phone",
                "voterid": "voter_id",
                "vote_status": "vote_status",
                "booth_number": "booth_number",
                "ward_number": "ward_number",
                "election_id": "election_id",
            }

            for source_key, target_key in aliases.items():
                if source_key in normalized and target_key not in normalized:
                    normalized[target_key] = normalized[source_key]

            return normalized

        if filename.endswith('.csv'):
            text_content = contents.decode('utf-8-sig')
            csv_file = io.StringIO(text_content)
            reader = csv.DictReader(csv_file)
            voter_rows = [normalize_row(row) for row in reader]
        elif filename.endswith('.xlsx'):
            from openpyxl import load_workbook

            workbook = load_workbook(io.BytesIO(contents), data_only=True)
            sheet = workbook.active
            excel_rows = list(sheet.iter_rows(values_only=True))

            if not excel_rows:
                raise HTTPException(status_code=400, detail='The uploaded Excel file is empty.')

            headers = [normalize_header(header) for header in excel_rows[0]]
            for row_values in excel_rows[1:]:
                if not row_values or not any(cell is not None and str(cell).strip() for cell in row_values):
                    continue

                row = {}
                for index, header in enumerate(headers):
                    if header:
                        row[header] = row_values[index] if index < len(row_values) else None
                voter_rows.append(normalize_row(row))
        else:
            raise HTTPException(status_code=400, detail='Only .csv and .xlsx files are supported. Please export the file to CSV or XLSX format.')

        if not voter_rows:
            raise HTTPException(status_code=400, detail='No voter rows were found in the uploaded file.')

        for row_number, row in enumerate(voter_rows, start=2):
            row['row_number'] = row_number
            row['name'] = str(row.get('name', '')).strip()
            row['aadhaar_number'] = str(row.get('aadhaar', '')).strip()
            row['state'] = str(row.get('state', '')).strip()
            row['phone'] = str(row.get('phone', '')).strip() or None
            row['voter_id'] = str(row.get('voter_id', '')).strip() or None
            row['voting_token'] = secrets.token_hex(32)
            row['voter_token'] = secrets.token_hex(32)
            row['registered_at'] = datetime.now().isoformat()

        if not hasattr(db, 'bulk_import_voters'):
            raise HTTPException(status_code=500, detail='Database backend does not support bulk import.')

        import_result = db.bulk_import_voters(voter_rows)
        total_rows = import_result.get('total', len(voter_rows))
        inserted = import_result.get('inserted', 0)
        updated = import_result.get('updated', 0)
        errors = import_result.get('errors', [])
        imported_count = inserted + updated

        return JSONResponse({
            'success': True,
            'message': f'Import complete: {imported_count} voters imported or updated, {len(errors)} errors found.',
            'total_rows': total_rows,
            'imported': imported_count,
            'imported_count': imported_count,
            'updated_count': updated,
            'error_count': len(errors),
            'errors': errors[:50]
        })
        
    except HTTPException:
        raise
    except Exception as e:
        import logging
        import traceback
        logging.error(f"Import error: {str(e)}")
        print(f"❌ Error importing voters: {str(e)}")
        print(f"❌ Full traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")


@app.get("/api/voters")
async def get_voters_deprecated(request: Request):
    """Get all voters (admin only) - deprecated, use /api/admin/voters."""
    session = check_admin_access(request)
    
    all_voters = db.get_all_voters()
    
    # Filter by state if not super admin
    admin_state = session.get("state")
    if admin_state != "All States":
        voters = [v for v in all_voters if v.get("state") == admin_state]
    else:
        voters = all_voters
    
    return {"voters": voters, "total": len(voters)}


@app.get("/api/votes")
async def get_all_votes(request: Request):
    """Get all votes (admin only)."""
    session = check_admin_access(request)
    
    elections = db.get_all_elections()
    
    # Filter by state if not super admin
    if session.get("state") != "All States":
        elections = [e for e in elections if e.get("state") == session.get("state")]
    
    all_votes = []
    for election in elections:
        votes = db.get_votes_by_election(election["id"])
        all_votes.extend(votes)
    
    return {"votes": all_votes}


@app.post("/api/vote")
async def cast_vote(vote: VoteRequest, request: Request):
    """Cast a vote - voters can only vote in their state's elections."""
    session = check_voter_access(request)
    
    print(f"🗳️ Vote request - Election ID: {vote.election_id}, Candidate ID: {vote.candidate_id}")
    
    if session.get("voter_token") != vote.voter_token:
        raise HTTPException(status_code=403, detail="Invalid voter token")
    
    election = db.get_election_by_id(vote.election_id)
    if not election:
        print(f"❌ Election not found: {vote.election_id}")
        raise HTTPException(status_code=404, detail="Election not found")
    
    # Verify voter is voting in their state's election
    voter_state = session.get("state")
    election_state = election.get("state")
    
    if voter_state != election_state:
        raise HTTPException(
            status_code=403, 
            detail=f"You can only vote in {voter_state} elections"
        )
    
    if election["status"] != "active":
        raise HTTPException(status_code=400, detail="Election is not active")
    
    if db.has_voted(vote.election_id, vote.voter_token):
        raise HTTPException(status_code=400, detail="You have already voted in this election")
    
    vote_data = {
        "election_id": vote.election_id,
        "candidate_id": vote.candidate_id,
        "voter_token": vote.voter_token,
        "state": voter_state,
        "timestamp": datetime.now().isoformat()
    }
    
    transaction_hash = db.record_vote(vote_data)
    
    return {
        "success": True,
        "transaction_hash": transaction_hash,
        "message": "Vote recorded successfully on blockchain"
    }


@app.get("/api/vote-status/{election_id}")
async def check_vote_status(election_id: str, request: Request):
    """Check if the current voter has already voted in an election."""
    session = check_voter_access(request)
    
    voter_token = session.get("voter_token")
    has_voted = db.has_voted(election_id, voter_token)
    
    return {
        "has_voted": has_voted,
        "election_id": election_id
    }


@app.post("/api/verify-vote")
async def verify_vote(verification: VoteVerification):
    """Verify a vote using transaction hash."""
    block = db.blockchain.get_block_by_hash(verification.transaction_hash)
    
    if not block:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    return {
        "success": True,
        "block": block,
        "verified": True,
        "message": "Vote verified on blockchain"
    }


@app.get("/api/elections/{election_id}/results")
async def get_results(election_id: str):
    """Get election results."""
    election = db.get_election_by_id(election_id)
    
    if not election:
        raise HTTPException(status_code=404, detail="Election not found")
    
    results = db.get_election_results(election_id)
    
    # Parse candidates if it's a JSON string
    candidates = election.get("candidates", [])
    if isinstance(candidates, str):
        import json
        try:
            candidates = json.loads(candidates)
        except:
            candidates = []
    
    candidates_with_votes = []
    for candidate in candidates:
        candidates_with_votes.append({
            **candidate,
            "votes": results.get(candidate["id"], 0)
        })
    
    total_votes = sum(results.values())
    
    return {
        "election": election,
        "results": candidates_with_votes,
        "total_votes": total_votes
    }


@app.get("/api/elections/{election_id}/export")
async def export_election_results(request: Request, election_id: str, format: str = "json"):
    """Export election results in JSON or CSV format (Admin only)."""
    session = check_admin_access(request)
    
    election = db.get_election_by_id(election_id)
    if not election:
        raise HTTPException(status_code=404, detail="Election not found")
    
    # Check if admin has access to this state's election
    admin_state = session.get("state")
    admin_role = session.get("role")
    
    if admin_role != "super_admin" and election["state"] != admin_state:
        raise HTTPException(
            status_code=403, 
            detail=f"Access denied. You can only export results for {admin_state} elections."
        )
    
    results = db.get_election_results(election_id)
    votes = db.get_votes_by_election(election_id)
    
    export_data = {
        "election_id": election_id,
        "title": election["title"],
        "state": election["state"],
        "description": election.get("description", ""),
        "status": election["status"],
        "start_time": election.get("start_time"),
        "end_time": election.get("end_time"),
        "total_votes": len(votes),
        "results": []
    }
    
    for candidate in election["candidates"]:
        vote_count = results.get(candidate["id"], 0)
        percentage = (vote_count / len(votes) * 100) if len(votes) > 0 else 0
        export_data["results"].append({
            "candidate_id": candidate["id"],
            "candidate_name": candidate["name"],
            "party": candidate["party"],
            "symbol": candidate.get("symbol", ""),
            "votes": vote_count,
            "percentage": round(percentage, 2)
        })
    
    # Sort by votes descending
    export_data["results"].sort(key=lambda x: x["votes"], reverse=True)
    
    if format.lower() == "csv":
        # Create CSV format
        from fastapi.responses import StreamingResponse
        import io
        import csv
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write election info
        writer.writerow(["Election Information"])
        writer.writerow(["Title", export_data["title"]])
        writer.writerow(["State", export_data["state"]])
        writer.writerow(["Status", export_data["status"]])
        writer.writerow(["Total Votes", export_data["total_votes"]])
        writer.writerow([])
        
        # Write results header
        writer.writerow(["Rank", "Candidate Name", "Party", "Symbol", "Votes", "Percentage"])
        
        # Write results data
        for idx, result in enumerate(export_data["results"], 1):
            writer.writerow([
                idx,
                result["candidate_name"],
                result["party"],
                result["symbol"],
                result["votes"],
                f"{result['percentage']}%"
            ])
        
        output.seek(0)
        filename = f"election_results_{election_id}_{election['state']}.csv"
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    else:
        # Return JSON format
        from fastapi.responses import JSONResponse
        filename = f"election_results_{election_id}_{election['state']}.json"
        
        return JSONResponse(
            content=export_data,
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )


@app.get("/api/blockchain")
async def get_blockchain():
    """Get the entire blockchain (for transparency)."""
    return {
        "chain": db.blockchain.get_chain(),
        "length": len(db.blockchain.chain),
        "is_valid": db.blockchain.is_chain_valid()
    }


@app.get("/api/blockchain/verify")
async def verify_blockchain():
    """Verify blockchain integrity."""
    is_valid = db.blockchain.is_chain_valid()
    
    return {
        "is_valid": is_valid,
        "chain_length": len(db.blockchain.chain),
        "message": "Blockchain is valid and tamper-proof" if is_valid else "Blockchain has been tampered with!"
    }


@app.get("/api/analytics/voter-turnout")
async def get_voter_turnout(request: Request, state: Optional[str] = None):
    """Get voter turnout analytics by state."""
    try:
        session = check_admin_access(request)
        admin_state = session.get("state")
        selected_state = admin_state

        if admin_state == "All States":
            if state and state != "All States" and state in INDIAN_STATES:
                selected_state = state
            else:
                selected_state = "All States"
        
        print(f"📊 Analytics request from admin state: {admin_state}, selected state: {selected_state}")
        
        # Get all voters
        all_voters = db.get_all_voters()
        print(f"📊 Total voters in system: {len(all_voters)}")
        
        # Get all elections to map election_id -> state
        all_elections = db.get_all_elections()
        elections_map = {e.get('id'): e.get('state') for e in all_elections if e.get('id')}

        # Get vote tracking to count unique voters who voted.
        # Prefer Supabase vote_tracking when available, otherwise derive from local votes.
        vote_tracking = []
        try:
            if hasattr(db, 'client'):
                vote_tracking_result = db.client.table('vote_tracking').select('*').execute()
                vote_tracking = vote_tracking_result.data if vote_tracking_result.data else []
                print(f"📊 Total vote tracking records: {len(vote_tracking)}")
        except Exception as e:
            print(f"⚠️ Error getting vote tracking: {e}")

        if not vote_tracking:
            all_votes = db.get_all_votes()
            for vote in all_votes:
                election_id = vote.get('election_id')
                # Local fallback stores voter_token instead of voter_token_hash.
                voter_identifier = vote.get('voter_token') or vote.get('voter_token_hash')
                if election_id and voter_identifier:
                    vote_tracking.append({
                        'election_id': election_id,
                        'voter_token_hash': voter_identifier
                    })
            print(f"📊 Derived vote tracking records from local votes: {len(vote_tracking)}")
        
        # Calculate statistics
        stats = {}
        
        if selected_state == "All States":
            # Super admin sees all states
            print(f"📊 Calculating stats for all states")
            for state in INDIAN_STATES:
                state_voters = [v for v in all_voters if v.get("state") == state]
                
                # Count UNIQUE voters from this state (not total votes)
                state_votes = [vt for vt in vote_tracking if elections_map.get(vt.get('election_id')) == state]
                # Get unique voter_token_hash values to count unique voters who voted
                unique_voters_who_voted = set(vt.get('voter_token_hash') for vt in state_votes if vt.get('voter_token_hash'))
                voted_count = len(unique_voters_who_voted)
                total_voters = len(state_voters)
                
                stats[state] = {
                    "total_voters": total_voters,
                    "voted_count": voted_count,
                    "turnout_percentage": round((voted_count / total_voters * 100), 2) if total_voters > 0 else 0
                }
                print(f"📊 {state}: {voted_count}/{total_voters} voters ({stats[state]['turnout_percentage']}%)")
        else:
            # State admin sees only their state, or super admin filtered to one state
            print(f"📊 Calculating stats for {selected_state}")
            state_voters = [v for v in all_voters if v.get("state") == selected_state]
            
            # Count UNIQUE voters from this state (not total votes)
            state_votes = [vt for vt in vote_tracking if elections_map.get(vt.get('election_id')) == selected_state]
            # Get unique voter_token_hash values to count unique voters who voted
            unique_voters_who_voted = set(vt.get('voter_token_hash') for vt in state_votes if vt.get('voter_token_hash'))
            voted_count = len(unique_voters_who_voted)
            total_voters = len(state_voters)
            
            stats[selected_state] = {
                "total_voters": total_voters,
                "voted_count": voted_count,
                "turnout_percentage": round((voted_count / total_voters * 100), 2) if total_voters > 0 else 0
            }
            print(f"📊 {selected_state}: {voted_count}/{total_voters} voters ({stats[selected_state]['turnout_percentage']}%)")
        
        return {"statistics": stats}
    
    except Exception as e:
        print(f"❌ Error in voter turnout analytics: {e}")
        import traceback
        traceback.print_exc()
        # Return empty stats instead of failing
        return {"statistics": {}}


@app.get("/api/analytics/participation-funnel")
async def get_participation_funnel(request: Request, state: Optional[str] = None):
    """Get a participation funnel for registered, contactable, authenticated, and voted voters."""
    try:
        session = check_admin_access(request)
        admin_state = session.get("state")
        selected_state = admin_state

        if admin_state == "All States" and state and state != "All States" and state in INDIAN_STATES:
            selected_state = state

        print(f"📊 Participation funnel request from admin state: {admin_state}, selected state: {selected_state}")

        all_voters = db.get_all_voters()
        all_elections = db.get_all_elections()
        elections_map = {e.get('id'): e.get('state') for e in all_elections if e.get('id')}
        vote_tracking = get_vote_tracking_records()

        if selected_state == "All States":
            relevant_voters = all_voters
            relevant_vote_tracking = vote_tracking
            relevant_sessions = db.get_all_sessions() if hasattr(db, 'get_all_sessions') else []
        else:
            relevant_voters = [v for v in all_voters if v.get('state') == selected_state]
            relevant_vote_tracking = [
                vt for vt in vote_tracking
                if elections_map.get(vt.get('election_id')) == selected_state
            ]
            relevant_sessions = [
                s for s in (db.get_all_sessions() if hasattr(db, 'get_all_sessions') else [])
                if s.get('type') == 'voter' and s.get('state') == selected_state
            ]

        registered_count = len(relevant_voters)
        contactable_count = len([
            voter for voter in relevant_voters
            if voter.get('phone') or voter.get('mobile')
        ])

        authenticated_ids = set()
        for voter_session in relevant_sessions:
            if voter_session.get('type') != 'voter':
                continue
            session_id = voter_session.get('user_id') or voter_session.get('token')
            if session_id:
                authenticated_ids.add(str(session_id))
        authenticated_count = len(authenticated_ids)

        voted_ids = set()
        for vote in relevant_vote_tracking:
            voter_identifier = vote.get('voter_token_hash') or vote.get('voter_token')
            if voter_identifier:
                voted_ids.add(str(voter_identifier))
        voted_count = len(voted_ids)

        def conversion_rate(current: int, previous: int) -> float:
            return round((current / previous * 100), 2) if previous > 0 else 0

        def drop_off(previous: int, current: int) -> int:
            return max(previous - current, 0)

        stages = [
            {
                "key": "registered",
                "label": "Registered Voters",
                "count": registered_count,
                "color": "#FF9933",
                "conversion_rate": 100
            },
            {
                "key": "contactable",
                "label": "Contactable Voters",
                "count": contactable_count,
                "color": "#000080",
                "conversion_rate": conversion_rate(contactable_count, registered_count)
            },
            {
                "key": "authenticated",
                "label": "Authenticated Sessions",
                "count": authenticated_count,
                "color": "#138808",
                "conversion_rate": conversion_rate(authenticated_count, contactable_count)
            },
            {
                "key": "voted",
                "label": "Votes Cast",
                "count": voted_count,
                "color": "#7c3aed",
                "conversion_rate": conversion_rate(voted_count, authenticated_count)
            }
        ]

        summary = {
            "selected_state": selected_state,
            "registered_count": registered_count,
            "contactable_count": contactable_count,
            "authenticated_count": authenticated_count,
            "voted_count": voted_count,
            "registration_to_contactable_dropoff": drop_off(registered_count, contactable_count),
            "contactable_to_authenticated_dropoff": drop_off(contactable_count, authenticated_count),
            "authenticated_to_voted_dropoff": drop_off(authenticated_count, voted_count),
            "final_conversion_rate": conversion_rate(voted_count, registered_count)
        }

        return {
            "state": selected_state,
            "summary": summary,
            "stages": stages
        }

    except Exception as e:
        print(f"❌ Error in participation funnel analytics: {e}")
        import traceback
        traceback.print_exc()
        return {"state": state or "All States", "summary": {}, "stages": []}


@app.get("/api/analytics/election-stats/{election_id}")
async def get_election_stats(election_id: str, request: Request):
    """Get detailed statistics for a specific election with live vote counts."""
    session = check_admin_access(request)
    
    election = db.get_election_by_id(election_id)
    if not election:
        raise HTTPException(status_code=404, detail="Election not found")
    
    # Verify admin has access to this election's state
    admin_state = session.get("state")
    if admin_state != "All States" and election["state"] != admin_state:
        raise HTTPException(status_code=403, detail="Access denied to this state's election")
    
    # Get vote counts (returns dict: candidate_id -> vote_count)
    vote_counts = db.get_election_results(election_id)
    
    # Parse candidates if it's a JSON string
    candidates = election.get("candidates", [])
    if isinstance(candidates, str):
        import json
        try:
            candidates = json.loads(candidates)
        except:
            candidates = []
    
    # Calculate statistics
    total_votes = sum(vote_counts.values())
    candidate_data = []
    
    for candidate in candidates:
        votes = vote_counts.get(candidate["id"], 0)
        percentage = round((votes / total_votes * 100), 2) if total_votes > 0 else 0
        candidate_data.append({
            "id": candidate.get("id"),
            "name": candidate.get("name", "Unknown"),
            "party": candidate.get("party", "Independent"),
            "symbol": candidate.get("symbol", ""),
            "photo": candidate.get("photo") or candidate.get("photo_url", ""),
            "logo": candidate.get("logo", ""),
            "votes": votes,
            "percentage": percentage
        })
    
    return {
        "election_id": election_id,
        "election_title": election["title"],
        "state": election["state"],
        "total_votes": total_votes,
        "candidates": candidate_data
    }


@app.post("/api/audit-log")
async def log_admin_action(request: Request, action: dict):
    """Log administrative actions for audit trail."""
    session = check_admin_access(request)
    action_type = (action or {}).get("type", "unknown")
    details = (action or {}).get("details", "")
    client_ip = request.client.host if request.client else "unknown"
    
    log_entry = {
        "username": session.get("username"),
        "action": f"{action_type}: {details}" if details else str(action_type),
        "action_type": action_type,
        "details": details,
        "state": session.get("state"),
        "role": session.get("role"),
        "ip_address": client_ip,
        "timestamp": datetime.now().isoformat()
    }
    
    db.save_audit_log(log_entry)
    
    return {"status": "logged"}


@app.get("/api/audit-logs")
async def get_audit_logs(
    request: Request,
    page: int = 1,
    page_size: int = 10,
    search: Optional[str] = None,
    action: Optional[str] = None,
    username: Optional[str] = None,
    state: Optional[str] = None,
    role: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None
):
    """Get audit logs (admin only)."""
    session = check_admin_access(request)
    admin_state = session.get("state")

    page = max(page, 1)
    page_size = max(min(page_size, 100), 5)

    raw_logs = db.get_audit_logs(limit=10000, offset=0)
    normalized_logs = [normalize_audit_log_entry(log) for log in raw_logs]

    if admin_state != "All States":
        normalized_logs = [log for log in normalized_logs if log.get("admin_state") == admin_state]

    filter_state = state if state and state != "All States" else None
    filters = {
        'search': search,
        'action': action,
        'username': username,
        'state': filter_state,
        'role': role,
        'from_date': from_date,
        'to_date': to_date
    }

    filtered_logs = [log for log in normalized_logs if audit_log_matches_filters(log, filters)]
    total = len(filtered_logs)
    total_pages = max((total + page_size - 1) // page_size, 1) if total else 0
    start_index = (page - 1) * page_size
    end_index = start_index + page_size
    page_logs = filtered_logs[start_index:end_index]

    return {
        "logs": page_logs,
        "page": page,
        "page_size": page_size,
        "total": total,
        "total_pages": total_pages
    }


@app.get("/api/verify-vote/{transaction_hash}")
async def verify_vote_endpoint(transaction_hash: str):
    """Public endpoint to verify a vote using its transaction hash."""
    try:
        print(f"🔍 Verifying transaction hash: {transaction_hash}")
        
        # Get the blockchain chain
        chain = db.blockchain.chain
        print(f"📊 Blockchain has {len(chain)} blocks")
        
        # Search through the blockchain for this transaction
        for i, block in enumerate(chain):
            # Convert Block object to dict
            block_dict = block.to_dict() if hasattr(block, 'to_dict') else block
            
            # Debug: print block info
            vote_data = block_dict.get("data", {})
            block_tx_hash = vote_data.get("transaction_hash", "")
            block_hash = block_dict.get("hash", "")
            
            print(f"Block {i}: tx_hash={block_tx_hash[:20]}..., block_hash={block_hash[:20]}...")
            
            # Check if transaction hash matches (check both locations)
            if block_tx_hash == transaction_hash or block_hash == transaction_hash:
                print(f"✅ Found matching block at index {i}")
                
                # Found the vote
                election_id = vote_data.get("election_id")
                
                if election_id:
                    election = db.get_election_by_id(election_id)
                    
                    return {
                        "verified": True,
                        "election_title": election.get("title", "Unknown") if election else "Unknown",
                        "state": election.get("state", "Unknown") if election else "Unknown",
                        "timestamp": block_dict.get("timestamp"),
                        "block_number": i,
                        "previous_hash": block_dict.get("previous_hash", "")[:16] + "...",
                        "current_hash": block_dict.get("hash", "")[:16] + "...",
                        "transaction_hash": block_tx_hash,
                        "message": "✅ Vote successfully verified on blockchain"
                    }
                else:
                    # Genesis block or non-vote block
                    return {
                        "verified": True,
                        "message": "✅ Block found in blockchain",
                        "block_number": i,
                        "timestamp": block_dict.get("timestamp"),
                        "current_hash": block_dict.get("hash", "")[:16] + "..."
                    }
        
        # Transaction not found - show what we have
        print(f"❌ Transaction hash not found in {len(chain)} blocks")
        print(f"Looking for: {transaction_hash}")
        
        # Debug: show all transaction hashes
        all_tx_hashes = []
        for block in chain:
            block_dict = block.to_dict() if hasattr(block, 'to_dict') else block
            tx_hash = block_dict.get("data", {}).get("transaction_hash", "N/A")
            all_tx_hashes.append(tx_hash[:20] + "...")
        print(f"Available tx hashes: {all_tx_hashes}")
        
        raise HTTPException(
            status_code=404,
            detail=f"Transaction hash not found. Blockchain has {len(chain)} blocks. Please verify the hash and try again."
        )
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error verifying vote: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error verifying vote: {str(e)}")


@app.get("/health")
async def health_check():
    """Health check endpoint."""
    return {"status": "healthy", "blockchain_valid": db.blockchain.is_chain_valid()}


@app.get("/api/public/statistics")
async def get_public_statistics():
    """Get public voting statistics - no authentication required."""
    try:
        elections = db.get_all_elections()
        votes = db.get_all_votes()
        
        # Calculate statistics by state
        state_stats = {}
        for state in INDIAN_STATES:
            state_stats[state] = {
                "total_elections": 0,
                "active_elections": 0,
                "completed_elections": 0,
                "total_votes": 0,
                "voter_turnout": 0
            }
        
        # Process elections
        for election in elections:
            state = election.get("state")
            if state in state_stats:
                state_stats[state]["total_elections"] += 1
                if election.get("status") == "active":
                    state_stats[state]["active_elections"] += 1
                elif election.get("status") == "ended":
                    state_stats[state]["completed_elections"] += 1
        
        # Process votes
        for vote in votes:
            election = db.get_election_by_id(vote.get("election_id"))
            if election:
                state = election.get("state")
                if state in state_stats:
                    state_stats[state]["total_votes"] += 1
        
        # Calculate overall statistics
        total_elections = len(elections)
        active_elections = sum(1 for e in elections if e.get("status") == "active")
        total_votes = len(votes)
        
        return {
            "success": True,
            "overall": {
                "total_elections": total_elections,
                "active_elections": active_elections,
                "total_votes": total_votes,
                "total_states": len([s for s, stats in state_stats.items() if stats["total_elections"] > 0])
            },
            "by_state": state_stats,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/admin/download-voter-template")
async def download_voter_template(request: Request):
    """
    Download sample voter data template (CSV format).
    Requires admin authentication.
    """
    from fastapi.responses import StreamingResponse
    
    # Check admin access
    check_admin_access(request)
    
    try:
        import csv
        import io

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=[
            'Name', 'Aadhaar', 'Phone', 'Voter_ID', 'Age', 'Gender',
            'City', 'State', 'Booth_Number', 'Ward_Number', 'Election_ID', 'Vote_Status'
        ])
        writer.writeheader()
        writer.writerow({
            'Name': 'Rajesh Kumar',
            'Aadhaar': '123456789012',
            'Phone': '919876543210',
            'Voter_ID': 'VOT1000',
            'Age': '45',
            'Gender': 'Male',
            'City': 'Mumbai',
            'State': 'Maharashtra',
            'Booth_Number': '12',
            'Ward_Number': '3',
            'Election_ID': 'ELEC100',
            'Vote_Status': 'Not Voted'
        })
        writer.writerow({
            'Name': 'Priya Sharma',
            'Aadhaar': '234567890123',
            'Phone': '919812345678',
            'Voter_ID': 'VOT1001',
            'Age': '39',
            'Gender': 'Female',
            'City': 'Delhi',
            'State': 'Delhi',
            'Booth_Number': '24',
            'Ward_Number': '8',
            'Election_ID': 'ELEC101',
            'Vote_Status': 'Voted'
        })
        writer.writerow({
            'Name': 'Amit Patel',
            'Aadhaar': '345678901234',
            'Phone': '919734567890',
            'Voter_ID': 'VOT1002',
            'Age': '52',
            'Gender': 'Male',
            'City': 'Ahmedabad',
            'State': 'Gujarat',
            'Booth_Number': '7',
            'Ward_Number': '21',
            'Election_ID': 'ELEC102',
            'Vote_Status': 'Not Voted'
        })

        output.seek(0)

        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            media_type="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=voter_import_template.csv"
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate template: {str(e)}")

# ──────────────────────────────────────────────────────────────────────────────
# NEW ENDPOINTS (added features)
# ──────────────────────────────────────────────────────────────────────────────

@app.get("/api/voter/profile")
async def get_voter_profile(request: Request):
    """Get the currently logged-in voter's profile and vote history."""
    session = check_voter_access(request)
    voter_id = session.get("voter_id")

    voter = db.get_voter(voter_id)
    if not voter:
        raise HTTPException(status_code=404, detail="Voter profile not found")

    # Get all elections to check voting history
    all_elections = db.get_all_elections()
    voter_token = session.get("voter_token", "")
    
    vote_history = []
    for election in all_elections:
        if db.has_voted(election["id"], voter_token):
            vote_history.append({
                "election_id": election["id"],
                "election_title": election.get("title", "Unknown"),
                "state": election.get("state", "Unknown"),
                "voted_at": election.get("created_at", "")
            })

    return {
        "success": True,
        "voter_id": voter_id,
        "name": voter.get("name", "Unknown"),
        "state": voter.get("state", "Unknown"),
        "registered_at": voter.get("registered_at", ""),
        "vote_count": len(vote_history),
        "vote_history": vote_history
    }


@app.patch("/api/admin/elections/{election_id}/status")
async def update_election_status(election_id: str, data: Dict, request: Request):
    """Manually update election status (admin only). Body: {"status": "active"|"ended"|"pending"}"""
    session = check_admin_access(request)
    admin_state = session.get("state")
    admin_role = session.get("role", "")

    election = db.get_election_by_id(election_id)
    if not election:
        raise HTTPException(status_code=404, detail="Election not found")

    # State-based access control
    if admin_role != "super_admin" and admin_state != "All States":
        if election.get("state") != admin_state:
            raise HTTPException(status_code=403, detail=f"Access denied. You can only manage {admin_state} elections.")

    new_status = data.get("status", "").lower()
    if new_status not in ("active", "ended", "pending"):
        raise HTTPException(status_code=400, detail="Invalid status. Must be 'active', 'ended', or 'pending'.")

    if hasattr(db, 'update_election_status'):
        db.update_election_status(election_id, new_status)
    else:
        # Fallback: update via save_election if update_election_status not available
        election["status"] = new_status
        db.save_election(election)

    print(f"[OK] Election {election_id} status updated to '{new_status}' by {session.get('username')}")
    return {
        "success": True,
        "election_id": election_id,
        "new_status": new_status,
        "message": f"Election status updated to '{new_status}'"
    }


@app.get("/api/public/live-results")
async def get_live_results(election_id: Optional[str] = None):
    """
    Public endpoint for real-time election results.
    Returns results for a specific election or all active elections.
    No authentication required.
    """
    try:
        if election_id:
            election = db.get_election_by_id(election_id)
            if not election:
                raise HTTPException(status_code=404, detail="Election not found")
            
            results = db.get_election_results(election_id)
            votes = db.get_votes_by_election(election_id)
            total_votes = len(votes)
            
            candidates = election.get("candidates", [])
            if isinstance(candidates, str):
                import json
                try:
                    candidates = json.loads(candidates)
                except Exception:
                    candidates = []

            candidate_results = []
            for c in candidates:
                vote_count = results.get(c["id"], 0)
                candidate_results.append({
                    "id": c["id"],
                    "name": c["name"],
                    "party": c.get("party", ""),
                    "symbol": c.get("symbol", ""),
                    "votes": vote_count,
                    "percentage": round((vote_count / total_votes * 100), 2) if total_votes > 0 else 0
                })
            candidate_results.sort(key=lambda x: x["votes"], reverse=True)

            return {
                "success": True,
                "election_id": election_id,
                "title": election.get("title"),
                "state": election.get("state"),
                "status": election.get("status"),
                "total_votes": total_votes,
                "results": candidate_results,
                "last_updated": datetime.now().isoformat()
            }
        else:
            # Return summary of all active elections
            elections = db.get_all_elections()
            active = [e for e in elections if e.get("status") == "active"]
            summaries = []
            for e in active:
                votes = db.get_votes_by_election(e["id"])
                summaries.append({
                    "election_id": e["id"],
                    "title": e.get("title"),
                    "state": e.get("state"),
                    "total_votes": len(votes),
                    "status": e.get("status")
                })
            return {
                "success": True,
                "active_elections": len(active),
                "elections": summaries,
                "last_updated": datetime.now().isoformat()
            }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/health/detailed")
async def detailed_health_check():
    """
    Detailed health check endpoint.
    Returns database status, blockchain integrity, and system info.
    """
    health = {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "database": {},
        "blockchain": {},
        "system": {}
    }

    # Check database
    try:
        elections = db.get_all_elections()
        voters = db.get_all_voters()
        db_type = type(db).__name__
        health["database"] = {
            "status": "connected",
            "type": db_type,
            "total_elections": len(elections),
            "total_voters": len(voters)
        }
    except Exception as e:
        health["database"] = {"status": "error", "error": str(e)}
        health["status"] = "degraded"

    # Check blockchain
    try:
        chain = db.blockchain.chain
        is_valid = db.blockchain.is_chain_valid()
        health["blockchain"] = {
            "status": "valid" if is_valid else "invalid",
            "chain_length": len(chain),
            "is_valid": is_valid
        }
        if not is_valid:
            health["status"] = "degraded"
    except Exception as e:
        health["blockchain"] = {"status": "error", "error": str(e)}

    # System info
    import platform
    health["system"] = {
        "python_version": sys.version.split()[0],
        "platform": platform.system(),
        "twilio_configured": TWILIO_AVAILABLE and bool(os.getenv("TWILIO_ACCOUNT_SID"))
    }

    return health


@app.get("/api/admin/system-info")
async def get_system_info(request: Request):
    """Get comprehensive system information for admins."""
    session = check_admin_access(request)
    admin_state = session.get("state")

    elections = db.get_all_elections()
    voters = db.get_all_voters()

    if admin_state != "All States":
        elections = [e for e in elections if e.get("state") == admin_state]
        voters = [v for v in voters if v.get("state") == admin_state]

    all_votes = []
    for e in elections:
        all_votes.extend(db.get_votes_by_election(e["id"]))

    active_elections = [e for e in elections if e.get("status") == "active"]
    ended_elections = [e for e in elections if e.get("status") == "ended"]

    return {
        "success": True,
        "admin_username": session.get("username"),
        "admin_state": admin_state,
        "admin_role": session.get("role"),
        "statistics": {
            "total_elections": len(elections),
            "active_elections": len(active_elections),
            "ended_elections": len(ended_elections),
            "total_voters": len(voters),
            "total_votes": len(all_votes),
            "voter_turnout_percent": round((len(all_votes) / len(voters) * 100), 2) if voters else 0
        },
        "blockchain": {
            "chain_length": len(db.blockchain.chain),
            "is_valid": db.blockchain.is_chain_valid()
        },
        "database_type": type(db).__name__,
        "timestamp": datetime.now().isoformat()
    }


@app.post("/api/admin/refresh-election-statuses")
async def refresh_election_statuses(request: Request):
    """Manually trigger election status auto-update (admin only)."""
    check_admin_access(request)
    try:
        _auto_update_election_statuses()
        return {"success": True, "message": "Election statuses refreshed successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn
    # Run on localhost only (not 0.0.0.0)
    uvicorn.run(app, host="localhost", port=5000, reload=False)

